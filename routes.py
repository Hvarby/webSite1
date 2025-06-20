from flask import render_template, redirect, url_for, request, abort, send_file
from flask_login import login_user, logout_user, login_required, current_user
from openpyxl import Workbook
from io import BytesIO
from forms.admin_edit_form import AdminEditForm
from models import db
from models.user import User
from models.project import Project
from models.role import Role
from models.log import Log
from models.user_project import UserProject
from forms.login_form import LoginForm
from forms.register_form import RegisterForm
from forms.edit_profile_form import EditProfileForm
from forms.assign_project_form import AssignProjectForm
from forms.assign_manager_form import AssignManagerForm
from log_action import log_action
from __init__ import bcrypt

def register_routes(app):

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        form = LoginForm()
        error = None
        if form.validate_on_submit():
            user = User.query.filter_by(email=form.email.data).first()
            if user and bcrypt.check_password_hash(user.password_hash, form.password.data):
                login_user(user)
                log_action("Вход в систему")
                return redirect(url_for('profile'))
            else:
                error = "Неверный логин или пароль."
        return render_template('login.html', form=form, error=error)

    @app.route('/logout')
    @login_required
    def logout():
        log_action("Выход из системы")
        logout_user()
        return redirect(url_for('login'))

    @app.route('/register', methods=['GET', 'POST'])
    def register():
        form = RegisterForm()
        if form.validate_on_submit():
            hashed_pw = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
            user = User(name=form.name.data, email=form.email.data, password_hash=hashed_pw, role_id=2)
            db.session.add(user)
            db.session.commit()
            return redirect(url_for('login'))
            log_action("Пользователь зарегестрирован")
        return render_template('register.html', form=form)

    @app.route('/profile')
    @login_required
    def profile():
        user_projects = [relation.project for relation in current_user.user_projects]
        return render_template('profile.html', user=current_user, projects=user_projects)

    @app.route('/edit_profile', methods=['GET', 'POST'])
    @login_required
    def edit_profile():
        form = EditProfileForm(email=current_user.email)

        if form.validate_on_submit():
            current_user.email = form.email.data

            # Если пользователь ввёл новый пароль — обновляем
            if form.password.data:
                hashed_pw = bcrypt.generate_password_hash(form.password.data).decode('utf-8')
                current_user.password_hash = hashed_pw

            db.session.commit()
            log_action("Изменение данных пользователя")
            return redirect(url_for('profile'))

        return render_template('edit_profile.html', form=form)

    @app.route('/projects')
    @login_required
    def projects():
        projects = Project.query.all()
        return render_template('projects.html', projects=projects)

    @app.route('/assign_project', methods=['GET', 'POST'])
    @login_required
    def assign_project():
        if current_user.role_id not in [1, 2, 3]:  # Только выше сотрудника
            return redirect(url_for('profile'))

        form = AssignProjectForm()
        form.user_id.choices = [
            (u.id, u.name) for u in User.query.filter(User.role_id > current_user.role_id).all()
        ]
        form.project_id.choices = [(p.id, p.title) for p in Project.query.all()]

        if form.validate_on_submit():
            # Проверяем, есть ли уже такая связь
            exists = UserProject.query.filter_by(
                user_id=form.user_id.data,
                project_id=form.project_id.data
            ).first()

            if exists:
                form.project_id.errors.append('Пользователь уже назначен на этот проект.')
            else:
                relation = UserProject(user_id=form.user_id.data, project_id=form.project_id.data)
                db.session.add(relation)
                db.session.commit()
                log_action("Назначил на проект пользователя")
                return redirect(url_for('assign_project'))
        return render_template('assign_project.html', form=form)

    @app.route('/users')
    @login_required
    def users():
        users = User.query.all()
        return render_template('users.html', users=users)

    @app.route('/admin/logs')
    @login_required
    def admin_logs():
        if current_user.role_id not in [1, 2, 3]:
            return redirect(url_for('profile'))

        page = request.args.get('page', 1, type=int)  # текущая страница
        per_page = 20  # количество логов на одной странице
        logs = Log.query.order_by(Log.timestamp.desc()).paginate(page=page, per_page=per_page)
        return render_template('admin_logs.html', logs=logs)

    @app.route('/admin/hierarchy', methods=['GET', 'POST'])
    @login_required
    def manage_hierarchy():
        if not current_user.role_id in  [1, 2, 3 ]:
            return redirect(url_for('profile'))

        form = AssignManagerForm()
        form.subordinate_id.choices = [(u.id, u.name) for u in User.query.filter(User.role_id > 1).all()]
        form.manager_id.choices = [(u.id, u.name) for u in User.query.filter(User.role_id < 4).all()]

        if form.validate_on_submit():
            subordinate = User.query.get(form.subordinate_id.data)
            subordinate.manager_id = form.manager_id.data
            db.session.commit()
            log_action("Назначил руководителя")
            return redirect(url_for('manage_hierarchy'))

        users = User.query.filter(User.role_id > 1).all()
        return render_template('hierarchy.html', users=users, form=form)

    @app.route('/admin/edit_user/<int:user_id>', methods=['GET', 'POST'])
    @login_required
    def edit_user_admin(user_id):
        if current_user.role_id != 1:
            return redirect(url_for('profile'))

        user = User.query.get_or_404(user_id)
        form = AdminEditForm(obj=user)
        form.role_id.choices = [(r.id, r.name) for r in Role.query.all()]

        if form.validate_on_submit():
            user.name = form.name.data
            user.email = form.email.data
            if user.id != current_user.id:
                user.role_id = form.role_id.data
            db.session.commit()
            log_action("Изменение данных пользователя Админом")
            return redirect(url_for('users'))
        return render_template('admin_edit_user.html', form=form, user=user)

    @app.route('/export')
    @login_required
    def export_options():
        if current_user.role_id not in [1, 2, 3]:
            return redirect(url_for('profile'))
        return render_template('export.html')

    @app.route('/export/projects')
    @login_required
    def export_projects():
        if current_user.role_id not in [1, 2, 3]:
            return redirect(url_for('profile'))
        log_action("Экспорт проектов")
        wb = Workbook()
        ws = wb.active
        ws.title = "Проекты"

        ws.append(["Название проекта", "Участник", "Роль", "Руководитель"])

        projects = Project.query.all()
        for project in projects:
            user_projects = UserProject.query.filter_by(project_id=project.id).all()
            if user_projects:
                for up in user_projects:
                    user = up.user
                    ws.append([
                        project.title,
                        user.name,
                        user.role.name,
                        user.manager.name if user.manager else "Нет руководителя"
                    ])
            else:
                ws.append([project.title, "Нет участников", "", ""])

        # Отправка файла
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            as_attachment=True,
            download_name="projects.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route('/export/users')
    @login_required
    def export_users():
        if current_user.role_id not in [1, 2, 3]:
            return redirect(url_for('profile'))
        log_action("Экспорт пользователей")
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи"

        ws.append(["Имя", "Email", "Роль", "Руководитель"])

        users = User.query.all()
        for user in users:
            ws.append([
                user.name,
                user.email,
                user.role.name,
                user.manager.name if user.manager else "Нет руководителя"
            ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            as_attachment=True,
            download_name="users.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route('/export/logs')
    @login_required
    def export_logs():
        if current_user.role_id not in [1, 2, 3]:
            return redirect(url_for('profile'))

        log_action("Экспорт логов")
        wb = Workbook()
        ws = wb.active
        ws.title = "Логи"
        ws.append(["Дата и время", "Пользователь", "Действие"])

        logs = Log.query.order_by(Log.timestamp.desc()).all()
        for log in logs:
            ws.append([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.user.name if log.user else "Неизвестно",
                log.action
            ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            as_attachment=True,
            download_name="logs.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


